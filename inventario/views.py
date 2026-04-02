from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Q
from coalfa.decorators import operacional_required
from .models import Lote, MovimientoTrazabilidad, Proveedor, Producto, RegistroServicio, RegistroTemperaturaCamara
from .forms import LoteForm
import datetime
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Helpers ────────────────────────────────────────────────────────────────────

def _build_arbol(lote, visitados=None):
    if visitados is None:
        visitados = set()
    if lote.pk in visitados:
        return []
    visitados.add(lote.pk)
    padres = []
    for padre in lote.lotes_padres.select_related("producto").all():
        padres.append({"lote": padre, "padres": _build_arbol(padre, visitados)})
    return padres


# ── Dashboard ──────────────────────────────────────────────────────────────────

@operacional_required
def dashboard_inventario(request):
    hoy = datetime.date.today()
    pronto = hoy + datetime.timedelta(days=7)
    
    # Caching para mejorar performance
    from django.core.cache import cache
    cache_key = "dashboard_inv_kpis"
    context = cache.get(cache_key)
    if context:
        return render(request, "inventario/dashboard.html", context)
    
    lotes_activos = Lote.objects.filter(estado="ACTIVO").select_related("producto", "proveedor")
    
    # KPIs de la empresa
    total_activos = lotes_activos.count()
    valor_total = sum(l.valor_total for l in lotes_activos)
    
    # Consolidamos las consultas agrupadas para evitar N+1 queries.
    from django.db.models import Sum, Count, ExpressionWrapper, F, FloatField
    
    # ── Alertas de Stock Crítico (Agrupados por Producto en una sola query) ──
    # Extraemos productos que tienen lotes activos y calculamos su stock actual.
    # Usamos annotate para calcular el stock total activo por producto directamente en la DB.
    productos_con_stock = Producto.objects.annotate(
        stock_actual=Sum('lotes__cantidad', filter=Q(lotes__estado='ACTIVO')),
        valor_total_p=Sum(ExpressionWrapper(F('lotes__cantidad') * F('lotes__precio_unitario'), output_field=FloatField()), filter=Q(lotes__estado='ACTIVO'))
    ).filter(lotes__isnull=False).distinct()

    productos_criticos = []
    cat_dist = {}
    cat_val_dist = {}

    for p in productos_con_stock:
        # Stock actual ya viene anotado. Si es None lo tratamos como 0.
        stock = p.stock_actual or 0
        valor = p.valor_total_p or 0
        
        # Lógica de Críticos
        if stock <= p.stock_minimo:
            if p.stock_minimo > 0 or (stock == 0 and p.lotes.filter(estado="CONSUMIDO").exists()):
                porcentaje = (stock / p.stock_minimo * 100) if p.stock_minimo > 0 else 0
                productos_criticos.append({
                    "producto": p,
                    "actual": stock,
                    "minimo": p.stock_minimo,
                    "porcentaje": porcentaje
                })
        
        # Distribución por Categoría (esto es eficiente hacerlo en memoria si no son miles)
        cat = p.get_categoria_display()
        cat_dist[cat] = cat_dist.get(cat, 0) + stock
        cat_val_dist[cat] = cat_val_dist.get(cat, 0) + valor
    
    cat_data = [{"label": k, "value": round(v, 2)} for k, v in cat_dist.items()]
    cat_val_data = [{"label": k, "value": round(v, 0)} for k, v in cat_val_dist.items()]

    # Alertas de Vencimiento
    alertas_vencimiento = Lote.objects.filter(
        fecha_vencimiento__lte=pronto, fecha_vencimiento__gte=hoy, estado="ACTIVO"
    ).order_by("fecha_vencimiento")
    
    vencidos_count = Lote.objects.filter(estado="ACTIVO", fecha_vencimiento__lt=hoy).count()

    # Timeline de movimientos (estilo empresarial)
    movimientos = MovimientoTrazabilidad.objects.select_related(
        "lote", "lote__producto"
    ).order_by("-fecha")[:12]

    # Kanban por ubicación optimizado (una sola query con Count)
    from django.db.models import Count
    conteos_por_area = lotes_activos.values("ubicacion_actual").annotate(count=Count("id"))
    counts_map = {row["ubicacion_actual"]: row["count"] for row in conteos_por_area}
    
    por_ubicacion = {}
    for key, label in Lote.UBICACION_CHOICES:
        por_ubicacion[key] = {
            "label": label,
            "count": counts_map.get(key, 0),
        }

    context = {
        "valor_total": valor_total,
        "total_activos": total_activos,
        "productos_criticos": sorted(productos_criticos, key=lambda x: x["porcentaje"])[:5],
        "criticos_count": len(productos_criticos),
        "vencidos_count": vencidos_count,
        "alertas": alertas_vencimiento.select_related("producto"),
        "movimientos": movimientos,
        "cat_data": cat_data,
        "cat_val_data": cat_val_data,
        "por_ubicacion": por_ubicacion,
        "max_ubicacion": max((d["count"] for d in por_ubicacion.values()), default=1) or 1,
        "hoy": hoy,
    }
    cache.set(cache_key, context, 600)  # 10 minutos
    return render(request, "inventario/dashboard.html", context)


@operacional_required
def mapa_bodega(request):
    """Vista visual que muestra el estado de las bodegas y cámaras."""
    hoy = datetime.date.today()
    proximo_vencimiento = hoy + datetime.timedelta(days=7)
    
    # Optimizamos a una sola consulta y agrupamos en memoria
    # Además pre-cargamos lotes que vencen pronto para detectar alertas una sola vez
    lotes_todos = Lote.objects.filter(estado="ACTIVO").select_related("producto")
    lotes_vencen_pronto = list(lotes_todos.filter(fecha_vencimiento__lte=proximo_vencimiento).values_list("ubicacion_actual", flat=True))
    
    # Agrupar en memoria para evitar consultas por sector
    sectores = {key: {"label": label, "lotes": [], "count": 0, "alerta": (key in lotes_vencen_pronto)} 
                for key, label in Lote.UBICACION_CHOICES}
                
    for lote in lotes_todos:
        if lote.ubicacion_actual in sectores:
            sectores[lote.ubicacion_actual]["lotes"].append(lote)
            sectores[lote.ubicacion_actual]["count"] += 1
    
    # Ultimas temperaturas de las 5 cámaras
    temperaturas = {cam_key: RegistroTemperaturaCamara.objects.filter(camara=cam_key).first() 
                    for cam_key, _ in RegistroTemperaturaCamara.CAMARA_CHOICES}

    context = {
        "sectores": sectores,
        "temperaturas": temperaturas,
        "hoy": hoy,
        "camara_choices": RegistroTemperaturaCamara.CAMARA_CHOICES,
    }
    return render(request, "inventario/mapa_bodega.html", context)


@operacional_required
def registrar_temperatura(request):
    """Registro rápido de temperatura para las cámaras."""
    if request.method == "POST":
        camara = request.POST.get("camara")
        temp = request.POST.get("temperatura")
        obs = request.POST.get("observaciones", "")
        
        if camara and temp:
            RegistroTemperaturaCamara.objects.create(
                camara=camara,
                temperatura=float(temp),
                usuario=request.user,
                observaciones=obs
            )
            messages.success(request, f"✅ Temperatura registrada para {camara}.")
        else:
            messages.error(request, "Faltan datos para el registro.")
            
    return redirect("inventario_mapa_bodega")




@operacional_required
def exportar_inventario_csv(request):
    q = request.GET.get("q", "")
    ubicacion = request.GET.get("ubicacion", "")
    estado = request.GET.get("estado", "")

    lotes = Lote.objects.select_related("producto", "proveedor").all()
    if q:
        lotes = lotes.filter(
            Q(numero_lote__icontains=q) | 
            Q(producto__nombre__icontains=q) |
            Q(numero_lote_proveedor__icontains=q)
        )
    if ubicacion:
        lotes = lotes.filter(ubicacion_actual=ubicacion)
    if estado:
        lotes = lotes.filter(estado=estado)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="inventario_{datetime.date.today()}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Código', 'Producto', 'Categoría', 'Cantidad', 'Unidad', 'Precio Unitario', 'Valor Total', 'Ubicación', 'Estado', 'Vencimiento'])
    
    for l in lotes:
        writer.writerow([
            l.numero_lote,
            l.producto.nombre,
            l.producto.get_categoria_display(),
            l.cantidad,
            l.producto.get_unidad_medida_display(),
            l.precio_unitario,
            l.valor_total,
            l.get_ubicacion_actual_display(),
            l.get_estado_display(),
            l.fecha_vencimiento
        ])
    
    return response


# ── Lista y Detalle de Lotes ───────────────────────────────────────────────────

@operacional_required
def lista_lotes(request):
    lotes = Lote.objects.select_related("producto", "proveedor")
    ubicacion = request.GET.get("ubicacion", "")
    estado = request.GET.get("estado", "")
    q = request.GET.get("q", "")
    if ubicacion:
        lotes = lotes.filter(ubicacion_actual=ubicacion)
    if estado:
        lotes = lotes.filter(estado=estado)
    if q:
        lotes = lotes.filter(numero_lote__icontains=q) | lotes.filter(
            producto__nombre__icontains=q
        )
    hoy = datetime.date.today()
    pronto = hoy + datetime.timedelta(days=7)
    context = {
        "lotes": lotes,
        "ubicacion": ubicacion,
        "estado": estado,
        "q": q,
        "ubicaciones": Lote.UBICACION_CHOICES,
        "estados": Lote.ESTADO_CHOICES,
        "hoy": hoy,
        "pronto": pronto,
    }
    return render(request, "inventario/lista_lotes.html", context)


@operacional_required
def detalle_lote(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    movimientos = lote.movimientos.all()
    destinos_posibles = Lote.UBICACION_CHOICES
    return render(request, "inventario/detalle_lote.html", {
        "lote": lote,
        "movimientos": movimientos,
        "destinos": destinos_posibles,
    })


@operacional_required
def editar_lote(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if request.method == "POST":
        form = LoteForm(request.POST, instance=lote)
        if form.is_valid():
            form.save()
            messages.success(request, f"Lote {lote.numero_lote} actualizado correctamente.")
            return redirect("inventario_detalle_lote", pk=pk)
    else:
        form = LoteForm(instance=lote)
    return render(request, "inventario/editar_lote.html", {"form": form, "lote": lote})


# ── Mover / Despachar Lote ─────────────────────────────────────────────────────

@operacional_required
def mover_lote(request, lote_id, nueva_ubicacion):
    lote = get_object_or_404(Lote, id=lote_id)
    if lote.estado != "ACTIVO":
        messages.error(request, "Solo se pueden mover lotes en estado Activo.")
        return redirect("inventario_detalle_lote", pk=lote_id)
    ubicaciones_validas = [k for k, _ in Lote.UBICACION_CHOICES]
    if nueva_ubicacion not in ubicaciones_validas:
        messages.error(request, "Ubicación no válida.")
        return redirect("inventario_detalle_lote", pk=lote_id)

    desde = lote.ubicacion_actual
    obs = request.POST.get("observaciones", "")

    # Cantidad a despachar (puede ser parcial)
    cantidad_str = request.POST.get("cantidad_despacho", "").strip()
    try:
        cantidad_despacho = float(cantidad_str) if cantidad_str else lote.cantidad
    except ValueError:
        messages.error(request, "Cantidad inválida.")
        return redirect("inventario_detalle_lote", pk=lote_id)

    if cantidad_despacho <= 0:
        messages.error(request, "La cantidad a despachar debe ser mayor a 0.")
        return redirect("inventario_detalle_lote", pk=lote_id)

    if cantidad_despacho > lote.cantidad:
        messages.error(request, f"No hay suficiente stock. Disponible: {lote.cantidad} {lote.producto.get_unidad_medida_display()}.")
        return redirect("inventario_detalle_lote", pk=lote_id)

    # Descontar stock y registrar movimiento
    MovimientoTrazabilidad.objects.create(
        lote=lote,
        desde=desde,
        hacia=nueva_ubicacion,
        cantidad=cantidad_despacho,
        responsable=request.user.get_full_name() or str(request.user),
        observaciones=obs,
    )

    # Si se mueve a un área de consumo/procesamiento, registrarlo como "Servicio" para seguimiento de costos (Task 1)
    areas_consumo = ["COCINA_FRIA", "COCINA_CALIENTE", "REPOSTERIA", "PANADERIA", "COLACION", "LINEA", "OTRO"]
    if nueva_ubicacion in areas_consumo and (desde not in areas_consumo or desde == nueva_ubicacion):
        RegistroServicio.objects.create(
            lote=lote,
            cantidad_servida=cantidad_despacho,
            area=nueva_ubicacion,
            responsable=request.user.get_full_name() or str(request.user),
            observaciones=obs if obs else (f"Despacho automático a {nueva_ubicacion}"),
        )
    lote.cantidad -= cantidad_despacho
    if lote.cantidad <= 0:
        lote.cantidad = 0
        lote.estado = "CONSUMIDO"
        messages.success(request, f"Stock agotado. Lote {lote.numero_lote} marcado como Consumido.")
    else:
        lote.ubicacion_actual = nueva_ubicacion
        messages.success(
            request,
            f"✅ {cantidad_despacho} {lote.producto.get_unidad_medida_display()} despachados a {nueva_ubicacion}. "
            f"Stock restante: {lote.cantidad} {lote.producto.get_unidad_medida_display()}."
        )
    lote.save()
    return redirect("inventario_detalle_lote", pk=lote_id)


# ── Eliminar Lote ──────────────────────────────────────────────────────────────

@operacional_required
def eliminar_lote(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if request.method == "POST":
        accion = request.POST.get("accion", "descartar")
        if accion == "eliminar":
            nombre = str(lote)
            lote.delete()
            messages.success(request, f"Lote {nombre} eliminado permanentemente.")
            return redirect("inventario_lista_lotes")
        else:
            lote.estado = "DESCARTADO"
            lote.save()
            messages.success(request, f"Lote {lote.numero_lote} marcado como Descartado.")
            return redirect("inventario_detalle_lote", pk=lote.pk)
    return render(request, "inventario/confirmar_eliminar.html", {
        "objeto": lote,
        "tipo": "lote",
        "nombre": str(lote),
        "advertencia": "Se eliminarán también todos los movimientos y servicios asociados." if request.GET.get("hard") else None,
    })


# ── Alertas ────────────────────────────────────────────────────────────────────

@operacional_required
def alertas_vencimiento(request):
    hoy = datetime.date.today()
    pronto = hoy + datetime.timedelta(days=7)
    vencidos = Lote.objects.filter(
        estado="ACTIVO", fecha_vencimiento__lt=hoy
    ).order_by("fecha_vencimiento")
    por_vencer = Lote.objects.filter(
        estado="ACTIVO",
        fecha_vencimiento__lte=pronto,
        fecha_vencimiento__gte=hoy,
    ).order_by("fecha_vencimiento")
    return render(request, "inventario/alertas.html", {
        "vencidos": vencidos, "por_vencer": por_vencer, "hoy": hoy,
    })


# ── Recepción ──────────────────────────────────────────────────────────────────

@operacional_required
def recibir_lote(request):
    productos = Producto.objects.all().order_by("nombre")
    proveedores = Proveedor.objects.all().order_by("nombre")
    unidades = Producto.UNIDAD_CHOICES

    if request.method == "POST":
        p = request.POST
        try:
            # Producto: buscar por nombre exacto (sin distinguir mayúsculas), o crear
            nombre_prod = p.get("producto_nombre", "").strip()
            talla_prod = p.get("talla", "").strip()
            unidad = p.get("unidad_medida", "KG")
            categoria = p.get("categoria", "OTRO")
            
            if not nombre_prod:
                messages.error(request, "Debes indicar el nombre del producto.")
                raise ValueError("Producto requerido")

            producto, creado_prod = Producto.objects.get_or_create(
                nombre__iexact=nombre_prod,
                talla__iexact=talla_prod,
                defaults={"nombre": nombre_prod, "talla": talla_prod, "unidad_medida": unidad, "categoria": categoria},
            )
            
            # Si el producto ya existía pero es la primera vez que le asignamos categoría
            if not creado_prod and producto.categoria == "OTRO" and categoria != "OTRO":
                producto.categoria = categoria
                producto.save()

            # Proveedor: buscar o crear
            nombre_prov = p.get("proveedor_nombre", "").strip()
            proveedor = None
            if nombre_prov:
                proveedor, _ = Proveedor.objects.get_or_create(
                    nombre__iexact=nombre_prov,
                    defaults={"nombre": nombre_prov},
                )

            lote = Lote(
                producto=producto,
                proveedor=proveedor,
                numero_lote_proveedor=p.get("numero_lote_proveedor", ""),
                numero_guia=p.get("numero_guia", ""),
                cantidad=float(p["cantidad"]),
                precio_unitario=float(p["precio_unitario"]) if p.get("precio_unitario") else 0,
                fecha_recepcion=p["fecha_recepcion"],
                fecha_vencimiento=p["fecha_vencimiento"],
                temperatura_recepcion=float(p["temperatura_recepcion"]) if p.get("temperatura_recepcion") else None,
                ubicacion_actual=p.get("ubicacion_actual", "BODEGA"),
                ubicacion_detalle=p.get("ubicacion_detalle", ""),
                proceso="RECEPCION",
                responsable_registro=request.user.get_full_name() or str(request.user),
                observaciones=p.get("observaciones", ""),
            )
            lote.save()
            
            nuevo = " (producto nuevo registrado)" if creado_prod else ""
            messages.success(request, f"✅ Lote {lote.numero_lote} registrado exitosamente.{nuevo}")
            return redirect("inventario_detalle_lote", pk=lote.pk)
        except Exception as e:
            messages.error(request, f"Error al guardar: {e}")

    return render(request, "inventario/recibir_lote.html", {
        "productos": productos,
        "proveedores": proveedores,
        "ubicaciones": Lote.UBICACION_CHOICES,
        "unidades": unidades,
        "categorias": Producto.CATEGORIA_CHOICES,
        "hoy": datetime.date.today().isoformat(),
    })


# ── Transformación ─────────────────────────────────────────────────────────────

@operacional_required
def transformar_lote(request):
    productos = Producto.objects.all().order_by("nombre")
    lotes_activos = Lote.objects.filter(estado="ACTIVO").select_related("producto").order_by("numero_lote")
    
    # — Sugerencia de Trazabilidad —
    # Lotes que fueron movidos a áreas de producción en las últimas 24 horas
    hace_24h = datetime.datetime.now() - datetime.timedelta(hours=24)
    movimientos_recientes = MovimientoTrazabilidad.objects.filter(
        fecha__gte=hace_24h,
        hacia__in=["REPOSTERIA", "COCINA_FRIA", "COCINA_CALIENTE", "LINEA"]
    ).values_list("lote_id", flat=True).distinct()
    
    id_sugeridos = list(movimientos_recientes)
    
    if request.method == "POST":
        p = request.POST
        lotes_padres_ids = request.POST.getlist("lotes_padres")
        if not lotes_padres_ids:
            messages.error(request, "Debes seleccionar al menos un lote insumo.")
        else:
            try:
                producto = Producto.objects.get(pk=p["producto"])
                lote_resultado = Lote(
                    producto=producto,
                    cantidad=float(p["cantidad"]),
                    fecha_recepcion=datetime.date.today(),
                    fecha_vencimiento=p["fecha_vencimiento"],
                    ubicacion_actual=p.get("ubicacion_destino", "COCINA_FRIA"),
                    proceso=p.get("proceso", "PREPARACION"),
                    responsable_registro=request.user.get_full_name() or str(request.user),
                    observaciones=p.get("observaciones", ""),
                )
                lote_resultado.save()
                padres = Lote.objects.filter(pk__in=lotes_padres_ids)
                lote_resultado.lotes_padres.set(padres)
                messages.success(request, f"✅ Lote {lote_resultado.numero_lote} creado con {len(lotes_padres_ids)} insumo(s).")
                return redirect("inventario_trazabilidad_lote", pk=lote_resultado.pk)
            except Exception as e:
                messages.error(request, f"Error al guardar: {e}")
    return render(request, "inventario/transformar_lote.html", {
        "productos": productos,
        "lotes_activos": lotes_activos,
        "id_sugeridos": id_sugeridos,
        "procesos": Lote.PROCESO_CHOICES[1:],
        "ubicaciones": Lote.UBICACION_CHOICES,
        "hoy": datetime.date.today().isoformat(),
    })


# ── Servicio ───────────────────────────────────────────────────────────────────

@operacional_required
def registrar_servicio(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if request.method == "POST":
        p = request.POST
        try:
            cantidad = float(p["cantidad_servida"])
            RegistroServicio.objects.create(
                lote=lote,
                cantidad_servida=cantidad,
                area=p.get("area", "LINEA"),
                responsable=request.user.get_full_name() or str(request.user),
                observaciones=p.get("observaciones", ""),
            )
            lote.cantidad -= cantidad
            if lote.cantidad <= 0:
                lote.cantidad = 0
                lote.estado = "CONSUMIDO"
            elif p.get("marcar_consumido"):
                lote.estado = "CONSUMIDO"
            lote.save()
            messages.success(request, f"✅ Servicio registrado para lote {lote.numero_lote}.")
            return redirect("inventario_detalle_lote", pk=lote.pk)
        except Exception as e:
            messages.error(request, f"Error: {e}")
    return render(request, "inventario/servicio_lote.html", {
        "lote": lote,
        "ubicaciones": Lote.UBICACION_CHOICES,
    })


# ── Trazabilidad ───────────────────────────────────────────────────────────────

@operacional_required
def trazabilidad_lote(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    arbol_padres = _build_arbol(lote)
    lotes_hijos = lote.lotes_hijos.select_related("producto").all()
    servicios = lote.servicios.all()
    movimientos = lote.movimientos.all()
    return render(request, "inventario/trazabilidad_lote.html", {
        "lote": lote,
        "arbol_padres": arbol_padres,
        "lotes_hijos": lotes_hijos,
        "servicios": servicios,
        "movimientos": movimientos,
    })


# ── Gestión de Productos ───────────────────────────────────────────────────────

@operacional_required
def lista_productos(request):
    from django.db.models import Sum, Count, Q
    
    # Anotamos stock total activo y conteo de lotes en una sola pasada por DB
    productos = Producto.objects.annotate(
        stock_total=Sum('lotes__cantidad', filter=Q(lotes__estado='ACTIVO')),
        cnt_lotes=Count('lotes', filter=Q(lotes__estado='ACTIVO'))
    ).order_by("nombre")
    
    data = []
    for p in productos:
        stock_total = p.stock_total or 0
        lotes_activos_cnt = p.cnt_lotes or 0
        
        # Cálculo de salud de stock
        salud = 100
        if p.stock_minimo > 0:
            salud = (stock_total / p.stock_minimo) * 100
        elif stock_total == 0:
            salud = 0
            
        data.append({
            "producto": p,
            "lotes_activos": lotes_activos_cnt,
            "stock_total": stock_total,
            "salud": min(salud, 100),  # Top at 100% for the bar
            "salud_real": salud,
        })
    return render(request, "inventario/lista_productos.html", {
        "data": data,
        "unidades": Producto.UNIDAD_CHOICES,
    })


@operacional_required
def eliminar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    total_lotes = producto.lotes.count()
    lotes_activos = producto.lotes.filter(estado="ACTIVO").count()
    bloqueado = total_lotes > 0  # PROTECT bloquea con cualquier lote, sin importar estado

    if request.method == "POST":
        if bloqueado:
            messages.error(request, f"No se puede eliminar '{producto.nombre}': tiene {total_lotes} lote(s) registrado(s) en el historial.")
            return redirect("inventario_lista_productos")
        nombre = producto.nombre
        producto.delete()
        messages.success(request, f"✅ Producto '{nombre}' eliminado correctamente.")
        return redirect("inventario_lista_productos")

    if total_lotes > 0:
        advertencia = (
            f"Este producto tiene {total_lotes} lote(s) en el historial "
            f"({lotes_activos} activo(s)) y NO puede eliminarse."
        )
    else:
        advertencia = None

    return render(request, "inventario/confirmar_eliminar.html", {
        "objeto": producto,
        "tipo": "producto",
        "nombre": producto.nombre,
        "advertencia": advertencia,
        "bloqueado": bloqueado,
    })




@operacional_required
def exportar_inventario_csv(request):
    q = request.GET.get("q", "")
    ubicacion = request.GET.get("ubicacion", "")
    estado = request.GET.get("estado", "")

    lotes = Lote.objects.select_related("producto", "proveedor").all()
    if q:
        lotes = lotes.filter(
            Q(numero_lote__icontains=q) |
            Q(producto__nombre__icontains=q) |
            Q(numero_lote_proveedor__icontains=q)
        )
    if ubicacion:
        lotes = lotes.filter(ubicacion_actual=ubicacion)
    if estado:
        lotes = lotes.filter(estado=estado)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="inventario_{datetime.date.today()}.csv"'
    writer = csv.writer(response)
    writer.writerow(["Codigo", "Producto", "Categoria", "Cantidad", "Unidad", "Precio Unitario", "Valor Total", "Ubicacion", "Estado", "Vencimiento"])
    for l in lotes:
        writer.writerow([
            l.numero_lote, l.producto.nombre, l.producto.get_categoria_display(),
            l.cantidad, l.producto.get_unidad_medida_display(),
            l.precio_unitario, l.valor_total,
            l.get_ubicacion_actual_display(), l.get_estado_display(), l.fecha_vencimiento
        ])
    return response


@operacional_required
def generar_reporte_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER
    import io

    hoy = datetime.date.today()
    lotes = Lote.objects.select_related("producto", "proveedor").filter(estado="ACTIVO").order_by("producto__nombre")
    valor_total = sum(l.valor_total for l in lotes)
    vencidos_pronto = Lote.objects.filter(
        estado="ACTIVO",
        fecha_vencimiento__lte=hoy + datetime.timedelta(days=7),
        fecha_vencimiento__gte=hoy
    ).count()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    ORANGE = colors.HexColor("#f97316")
    DARK   = colors.HexColor("#0b1120")
    GRAY   = colors.HexColor("#64748b")
    LIGHT  = colors.HexColor("#f8fafc")
    DANGER = colors.HexColor("#f87171")
    GREEN  = colors.HexColor("#34d399")

    title_s  = ParagraphStyle("title",  fontSize=22, textColor=ORANGE, fontName="Helvetica-Bold", spaceAfter=2)
    sub_s    = ParagraphStyle("sub",    fontSize=10, textColor=GRAY,   fontName="Helvetica", spaceAfter=4)
    cell_s   = ParagraphStyle("cell",   fontSize=8,  fontName="Helvetica", textColor=LIGHT)
    head_s   = ParagraphStyle("head",   fontSize=8,  fontName="Helvetica-Bold", textColor=colors.white)
    footer_s = ParagraphStyle("footer", fontSize=7,  textColor=GRAY, alignment=TA_CENTER)

    elements = []

    elements.append(Paragraph("Centro Medico San Lucas", sub_s))
    elements.append(Paragraph("Reporte de Inventario Activo", title_s))
    elements.append(Paragraph(f"Generado: {hoy.strftime('%d/%m/%Y')}   Lotes: {lotes.count()}", sub_s))
    elements.append(HRFlowable(width="100%", thickness=2, color=ORANGE, spaceAfter=12))

    # KPI panel
    kpi_h = [Paragraph(x, head_s) for x in ["LOTES ACTIVOS", "VALOR TOTAL", "POR VENCER (7d)", "FECHA"]]
    kpi_v = [
        Paragraph(str(lotes.count()), ParagraphStyle("kv1", fontSize=18, fontName="Helvetica-Bold", textColor=ORANGE)),
        Paragraph(f"${valor_total:,.0f}", ParagraphStyle("kv2", fontSize=18, fontName="Helvetica-Bold", textColor=GREEN)),
        Paragraph(str(vencidos_pronto), ParagraphStyle("kv3", fontSize=18, fontName="Helvetica-Bold", textColor=DANGER)),
        Paragraph(hoy.strftime("%d/%m/%Y"), ParagraphStyle("kv4", fontSize=18, fontName="Helvetica-Bold", textColor=LIGHT)),
    ]
    kpi_table = Table([kpi_h, kpi_v], colWidths=["25%", "25%", "25%", "25%"])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), DARK),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#10192d")),
        ("BOX", (0, 0), (-1, -1), 1, ORANGE),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#1e293b")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 16))

    # Main inventory table
    headers = [Paragraph(x, head_s) for x in
               ["CODIGO", "PRODUCTO", "PROVEEDOR", "CANTIDAD", "PRECIO UNIT.", "VALOR TOTAL", "AREA", "VENCIMIENTO"]]
    table_data = [headers]
    for l in lotes:
        dias = l.dias_restantes
        vc = DANGER if dias <= 0 else (colors.HexColor("#fbbf24") if dias <= 7 else GREEN)
        table_data.append([
            Paragraph(l.numero_lote, ParagraphStyle("code", fontSize=8, fontName="Helvetica-Bold", textColor=ORANGE)),
            Paragraph(l.producto.nombre, cell_s),
            Paragraph(l.proveedor.nombre if l.proveedor else "-", cell_s),
            Paragraph(f"{l.cantidad} {l.producto.get_unidad_medida_display()}", cell_s),
            Paragraph(f"${l.precio_unitario:,.0f}", cell_s),
            Paragraph(f"${l.valor_total:,.0f}", ParagraphStyle("vt", fontSize=8, fontName="Helvetica-Bold", textColor=GREEN)),
            Paragraph(l.get_ubicacion_actual_display(), cell_s),
            Paragraph(l.fecha_vencimiento.strftime("%d/%m/%Y"), ParagraphStyle("date", fontSize=8, fontName="Helvetica-Bold", textColor=vc)),
        ])

    main_t = Table(table_data, colWidths=[3.5*cm, 5*cm, 4*cm, 2.5*cm, 2.5*cm, 2.8*cm, 3.5*cm, 2.8*cm], repeatRows=1)
    main_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), DARK),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#1e293b")),
        ("BOX", (0, 0), (-1, -1), 0.5, ORANGE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#0b1120"), colors.HexColor("#10192d")]),
    ]))
    elements.append(main_t)
    elements.append(Spacer(1, 12))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=GRAY))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"Reporte confidencial - Centro Medico San Lucas - {hoy.strftime('%Y')}", footer_s))

    doc.build(elements)
    buffer.seek(0)
    resp = HttpResponse(buffer, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="reporte_inventario_{hoy}.pdf"'
    return resp
@operacional_required
def exportar_inventario_excel(request):
    """Exporta el inventario a Excel con formato profesional empresarial."""
    q = request.GET.get("q", "")
    ubicacion = request.GET.get("ubicacion", "")
    estado = request.GET.get("estado", "ACTIVO")

    lotes = Lote.objects.select_related("producto", "proveedor").all()
    if q:
        lotes = lotes.filter(
            Q(numero_lote__icontains=q) |
            Q(producto__nombre__icontains=q) |
            Q(numero_lote_proveedor__icontains=q)
        )
    if ubicacion:
        lotes = lotes.filter(ubicacion_actual=ubicacion)
    if estado:
        lotes = lotes.filter(estado=estado)

    hoy = datetime.date.today()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # ─── Estilos ───
    color_naranja  = "F97316"
    color_oscuro   = "0B1120"
    color_gris     = "64748B"
    color_verde    = "22C55E"
    color_rojo     = "EF4444"
    color_amarillo = "EAB308"
    color_fila_a   = "111827"
    color_fila_b   = "1F2937"

    font_titulo  = Font(name="Calibri", size=16, bold=True, color=color_naranja)
    font_empresa = Font(name="Calibri", size=10, color=color_gris, italic=True)
    font_header  = Font(name="Calibri", size=9,  bold=True, color="FFFFFF")
    font_normal  = Font(name="Calibri", size=9,  color="F1F5F9")
    font_codigo  = Font(name="Courier New", size=8, bold=True, color=color_naranja)
    font_total   = Font(name="Calibri", size=10, bold=True, color=color_verde)
    font_muted   = Font(name="Calibri", size=8, color="94A3B8")

    fill_titulo  = PatternFill("solid", fgColor=color_oscuro)
    fill_header  = PatternFill("solid", fgColor="1E293B")
    fill_fila_a  = PatternFill("solid", fgColor=color_fila_a)
    fill_fila_b  = PatternFill("solid", fgColor=color_fila_b)
    fill_total   = PatternFill("solid", fgColor="0F2027")

    thin = Side(style="thin", color="374151")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    borde_total = Border(left=thin, right=thin, top=Side(style="medium", color=color_naranja), bottom=thin)

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right",  vertical="center")

    # ─── Fila 1: Título empresa ───
    ws.merge_cells("A1:J1")
    ws["A1"] = "CENTRO MÉDICO SAN LUCAS"
    ws["A1"].font = font_empresa
    ws["A1"].fill = fill_titulo
    ws["A1"].alignment = left
    ws.row_dimensions[1].height = 18

    # ─── Fila 2: Título reporte ───
    ws.merge_cells("A2:J2")
    ws["A2"] = "REPORTE DE INVENTARIO"
    ws["A2"].font = font_titulo
    ws["A2"].fill = fill_titulo
    ws["A2"].alignment = left
    ws.row_dimensions[2].height = 28

    # ─── Fila 3: Metadatos ───
    ws.merge_cells("A3:E3")
    ws["A3"] = f"Generado el {hoy.strftime('%d/%m/%Y')}   |   Filtro estado: {estado or 'Todos'}"
    ws["A3"].font = font_muted
    ws["A3"].fill = fill_titulo
    ws["A3"].alignment = left
    ws.merge_cells("F3:J3")
    valor_total = sum(l.valor_total for l in lotes)
    ws["F3"] = f"Valor Total: ${valor_total:,.0f}   |   Lotes: {lotes.count()}"
    ws["F3"].font = Font(name="Calibri", size=9, bold=True, color=color_verde)
    ws["F3"].fill = fill_titulo
    ws["F3"].alignment = right
    ws.row_dimensions[3].height = 16

    # ─── Fila 4: Vacía (separador) ───
    ws.merge_cells("A4:J4")
    ws["A4"].fill = fill_titulo
    ws.row_dimensions[4].height = 6

    # ─── Fila 5: Encabezados ───
    HEADERS = [
        "Código Lote", "Producto", "Categoría", "Proveedor",
        "Cantidad", "Unidad", "Precio Unit.", "Valor Total",
        "Área / Ubicación", "Vencimiento"
    ]
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = center
        cell.border = borde
    ws.row_dimensions[5].height = 22

    # ─── Columnas de anchura ───
    COL_WIDTHS = [18, 28, 16, 22, 10, 10, 14, 14, 22, 14]
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ─── Datos ───
    total_valor = 0
    for row_num, lote in enumerate(lotes, start=6):
        fill = fill_fila_a if row_num % 2 == 0 else fill_fila_b
        dias = lote.dias_restantes
        color_venc = color_rojo if dias <= 0 else (color_amarillo if dias <= 7 else color_verde)
        venc_font = Font(name="Calibri", size=9, bold=True, color=color_venc)

        ubicacion_label = lote.get_ubicacion_actual_display()
        if lote.ubicacion_actual == "OTRO" and lote.ubicacion_detalle:
            ubicacion_label = f"Otro: {lote.ubicacion_detalle}"
        elif lote.ubicacion_detalle:
            ubicacion_label = f"{ubicacion_label} — {lote.ubicacion_detalle}"

        fila = [
            (lote.numero_lote,          font_codigo,   left),
            (lote.producto.nombre,      font_normal,   left),
            (lote.producto.get_categoria_display(), font_normal, center),
            (lote.proveedor.nombre if lote.proveedor else "-", font_normal, left),
            (lote.cantidad,             font_normal,   right),
            (lote.producto.get_unidad_medida_display(), font_normal, center),
            (lote.precio_unitario,      font_normal,   right),
            (lote.valor_total,          Font(name="Calibri", size=9, bold=True, color=color_verde), right),
            (ubicacion_label,           font_normal,   left),
            (lote.fecha_vencimiento.strftime("%d/%m/%Y"), venc_font, center),
        ]
        total_valor += lote.valor_total

        for col_idx, (value, font_, align_) in enumerate(fila, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = font_
            cell.fill = fill
            cell.alignment = align_
            cell.border = borde
            # Formato numérico
            if col_idx in (7, 8):  # precio y valor
                cell.number_format = '$#,##0.00'
            elif col_idx == 5:     # cantidad
                cell.number_format = '#,##0.00'
        ws.row_dimensions[row_num].height = 18

    # ─── Fila de totales ───
    tot_row = lotes.count() + 6
    ws.merge_cells(f"A{tot_row}:G{tot_row}")
    ws[f"A{tot_row}"] = "TOTAL VALOR INVENTARIO"
    ws[f"A{tot_row}"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws[f"A{tot_row}"].fill = fill_total
    ws[f"A{tot_row}"].alignment = right
    ws[f"A{tot_row}"].border = borde_total
    ws[f"H{tot_row}"] = total_valor
    ws[f"H{tot_row}"].font = font_total
    ws[f"H{tot_row}"].fill = fill_total
    ws[f"H{tot_row}"].alignment = right
    ws[f"H{tot_row}"].border = borde_total
    ws[f"H{tot_row}"].number_format = '$#,##0.00'
    for c in ["I", "J"]:
        ws[f"{c}{tot_row}"].fill = fill_total
        ws[f"{c}{tot_row}"].border = borde_total
    ws.row_dimensions[tot_row].height = 22

    # ─── Inmovilizar encabezados ───
    ws.freeze_panes = "A6"

    # ─── Filtros automáticos ───
    ws.auto_filter.ref = f"A5:J{tot_row - 1}"

    # ─── Fondo oscuro general ───
    for row in ws.iter_rows(min_row=1, max_row=4):
        for cell in row:
            if not cell.fill or cell.fill.fgColor.value == "00000000":
                cell.fill = fill_titulo

    # ─── Respuesta HTTP ───
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="Inventario_{hoy}.xlsx"'
    wb.save(response)
    return response

@operacional_required
def resumen_consumo(request):
    """Muestra el ranking de áreas que más consumen (Task 4)."""
    # Ranking de áreas por gasto monetario
    ranking = RegistroServicio.objects.values('area').annotate(
        total_gastado=Sum('costo_total'),
        total_productos=Sum('cantidad_servida')
    ).order_by('-total_gastado')
    
    # Convertir keys de area a labels (human-readable)
    ubicaciones_dict = dict(Lote.UBICACION_CHOICES)
    for r in ranking:
        r['area_label'] = ubicaciones_dict.get(r['area'], r['area'])

    # Últimos 20 registros de consumo
    ultimos_consumos = RegistroServicio.objects.select_related('lote', 'lote__producto').order_by('-fecha')[:20]

    return render(request, "inventario/resumen_consumo.html", {
        "ranking": ranking,
        "ultimos_consumos": ultimos_consumos,
        "ubicaciones": Lote.UBICACION_CHOICES,
    })

@operacional_required
def exportar_consumo_excel(request):
    """Genera planilla Excel de consumo por área (Task 2) con formato profesional y totales."""
    desde = request.GET.get("desde")
    hasta = request.GET.get("hasta")
    area_filtro = request.GET.get("area")
    periodo = request.GET.get("periodo")

    hoy = datetime.date.today()
    
    # Lógica de períodos rápidos
    if periodo == "semanal":
        desde = (hoy - datetime.timedelta(days=7)).isoformat()
        hasta = hoy.isoformat()
    elif periodo == "mensual":
        # Desde el primer día del mes actual
        desde = hoy.replace(day=1).isoformat()
        hasta = hoy.isoformat()

    consumos = RegistroServicio.objects.select_related('lote', 'lote__producto').all()
    if desde:
        consumos = consumos.filter(fecha__date__gte=desde)
    if hasta:
        consumos = consumos.filter(fecha__date__lte=hasta)
    if area_filtro:
        consumos = consumos.filter(area=area_filtro)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Consumo"

    # ─── Estilos (Consistentes con el resto del sistema) ───
    color_naranja  = "F97316"
    color_oscuro   = "0B1120"
    color_gris     = "64748B"
    color_verde    = "22C55E"
    color_fila_a   = "111827"
    color_fila_b   = "1F2937"

    font_titulo  = Font(name="Calibri", size=16, bold=True, color=color_naranja)
    font_empresa = Font(name="Calibri", size=10, color=color_gris, italic=True)
    font_header  = Font(name="Calibri", size=9,  bold=True, color="FFFFFF")
    font_normal  = Font(name="Calibri", size=9,  color="F1F5F9")
    font_total   = Font(name="Calibri", size=10, bold=True, color=color_verde)
    font_muted   = Font(name="Calibri", size=8, color="94A3B8")

    fill_titulo  = PatternFill("solid", fgColor=color_oscuro)
    fill_header  = PatternFill("solid", fgColor="1E293B")
    fill_fila_a  = PatternFill("solid", fgColor=color_fila_a)
    fill_fila_b  = PatternFill("solid", fgColor=color_fila_b)
    fill_total   = PatternFill("solid", fgColor="0F2027")

    thin = Side(style="thin", color="374151")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    borde_total = Border(left=thin, right=thin, top=Side(style="medium", color=color_naranja), bottom=thin)

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right",  vertical="center")

    # ─── Encabezado de Reporte ───
    ws.merge_cells("A1:I1")
    ws["A1"] = "CENTRO MÉDICO SAN LUCAS"
    ws["A1"].font = font_empresa
    ws["A1"].fill = fill_titulo
    ws["A1"].alignment = left
    
    ws.merge_cells("A2:I2")
    tit = "REPORTE DE CONSUMO"
    if periodo: tit += f" ({periodo.upper()})"
    ws["A2"] = tit
    ws["A2"].font = font_titulo
    ws["A2"].fill = fill_titulo
    ws["A2"].alignment = left

    ws.merge_cells("A3:I3")
    filtro_info = f"Desde: {desde or 'Inicio'} | Hasta: {hasta or 'Hoy'}"
    if area_filtro: filtro_info += f" | Área: {area_filtro}"
    ws["A3"] = f"Generado el {hoy.strftime('%d/%m/%Y')}   |   {filtro_info}"
    ws["A3"].font = font_muted
    ws["A3"].fill = fill_titulo
    ws["A3"].alignment = left
    
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 16

    # ─── Espacio ───
    for cell in ws["A4:I4"][0]: cell.fill = fill_titulo
    ws.row_dimensions[4].height = 6

    # ─── Tabla ───
    headers = ["Fecha", "Área", "Producto", "Código Lote", "Cantidad", "Unidad", "Costo Unit.", "Costo Total", "Responsable"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = center
        cell.border = borde
    ws.row_dimensions[5].height = 22

    # Column Widths
    COL_WIDTHS = [20, 18, 28, 18, 12, 10, 14, 14, 20]
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ─── Datos ───
    total_cant = 0
    total_costo = 0
    row_num = 6
    for c in consumos:
        fill = fill_fila_a if row_num % 2 == 0 else fill_fila_b
        
        fila_data = [
            (c.fecha.strftime("%d/%m/%Y %H:%M"), font_normal, center),
            (c.get_area_display(),               font_normal, center),
            (c.lote.producto.nombre,             font_normal, left),
            (c.lote.numero_lote,                 font_normal, left),
            (c.cantidad_servida,                 font_normal, right),
            (c.lote.producto.get_unidad_medida_display(), font_normal, center),
            (c.lote.precio_unitario,             font_normal, right),
            (c.costo_total,                      Font(name="Calibri", size=9, bold=True, color=color_verde), right),
            (c.responsable,                      font_normal, left),
        ]
        
        total_cant += c.cantidad_servida
        total_costo += c.costo_total

        for col_idx, (value, font_, align_) in enumerate(fila_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = font_
            cell.fill = fill
            cell.alignment = align_
            cell.border = borde
            if col_idx in (7, 8): cell.number_format = "$#,##0"
            elif col_idx == 5:    cell.number_format = "#,##0.00"

        ws.row_dimensions[row_num].height = 18
        row_num += 1

    # ─── Fila de Totales (REQUERIMIENTO 1) ───
    ws.merge_cells(f"A{row_num}:D{row_num}")
    ws[f"A{row_num}"] = "TOTALES GENERALES"
    ws[f"A{row_num}"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws[f"A{row_num}"].fill = fill_total
    ws[f"A{row_num}"].alignment = right
    ws[f"A{row_num}"].border = borde_total

    # Cantidad Total
    ws[f"E{row_num}"] = total_cant
    ws[f"E{row_num}"].font = font_total
    ws[f"E{row_num}"].fill = fill_total
    ws[f"E{row_num}"].alignment = right
    ws[f"E{row_num}"].border = borde_total
    ws[f"E{row_num}"].number_format = "#,##0.00"

    # Costo Total
    ws[f"H{row_num}"] = total_costo
    ws[f"H{row_num}"].font = font_total
    ws[f"H{row_num}"].fill = fill_total
    ws[f"H{row_num}"].alignment = right
    ws[f"H{row_num}"].border = borde_total
    ws[f"H{row_num}"].number_format = "$#,##0"

    # Celdas vacías del total para el borde/fondo
    for col in ["F", "G", "I"]:
        ws[f"{col}{row_num}"].fill = fill_total
        ws[f"{col}{row_num}"].border = borde_total

    ws.row_dimensions[row_num].height = 24

    # Freeze and Filter
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:I{row_num - 1}"

    # General Dark Background for unused top cells
    for r in ws.iter_rows(min_row=1, max_row=4):
        for cell in r:
            if not cell.fill or cell.fill.fgColor.value == "00000000":
                cell.fill = fill_titulo

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Consumo_Bodega_{hoy}.xlsx"'
    wb.save(response)
    return response
