from django.shortcuts import render, get_object_or_404, redirect
from django.db import models
from django.db.models import Sum
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from coalfa.decorators import rrhh_required
from .models import Empleado, PeriodoAusencia, Documento, GastoRRHH, ProductoEPP, EntregaEPP
from .forms import EmpleadoForm, DocumentoEditForm, GastoRRHHForm, EntregaEPPForm
from django import forms
import datetime
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


class SubirDocumentoForm(forms.ModelForm):
    class Meta:
        model = Documento
        fields = ("tipo", "descripcion", "archivo")
        widgets = {
            "tipo": forms.Select(attrs={"class": "form-control"}),
            "descripcion": forms.TextInput(attrs={"class": "form-control", "placeholder": "Descripción breve (opcional, ej: Documento cargado masivamente)"}),
            "archivo": forms.FileInput(attrs={"class": "form-control"}),
        }


@rrhh_required
def dashboard_rrhh(request):
    total = Empleado.objects.count()
    activos = Empleado.objects.filter(estado="ACTIVO").count()
    vacaciones = Empleado.objects.filter(estado="VACACIONES").count()
    licencia = Empleado.objects.filter(estado="LICENCIA").count()
    finiquitados = Empleado.objects.filter(estado="FINIQUITADO").count()
    hoy = datetime.date.today()
    ausencias_activas = PeriodoAusencia.objects.filter(fecha_inicio__lte=hoy, fecha_fin__gte=hoy)
    context = {
        "total": total, "activos": activos, "vacaciones": vacaciones,
        "licencia": licencia, "finiquitados": finiquitados,
        "ausencias_activas": ausencias_activas[:5],
    }
    return render(request, "rrhh/dashboard.html", context)


@rrhh_required
def lista_empleados(request):
    empleados = Empleado.objects.all()
    estado = request.GET.get("estado", "")
    area = request.GET.get("area", "")
    q = request.GET.get("q", "")
    if estado:
        empleados = empleados.filter(estado=estado)
    if area:
        empleados = empleados.filter(area=area)
    if q:
        empleados = empleados.filter(nombre__icontains=q) | empleados.filter(apellido__icontains=q) | empleados.filter(rut__icontains=q)
    
    # Form for the "New Employee" modal
    nuevo_form = EmpleadoForm()
    if request.method == "POST" and "crear_empleado" in request.POST:
        nuevo_form = EmpleadoForm(request.POST, request.FILES)
        if nuevo_form.is_valid():
            emp = nuevo_form.save()
            messages.success(request, f"✅ Empleado {emp.nombre} {emp.apellido} registrado correctamente.")
            return redirect("rrhh_lista_empleados")
        # If invalid, fall through to render with errors

    context = {
        "empleados": empleados, "estado": estado, "area": area, "q": q,
        "estados": Empleado.ESTADO_CHOICES, "areas": Empleado.AREA_CHOICES,
        "nuevo_form": nuevo_form,
        "show_modal": (request.method == "POST" and "crear_empleado" in request.POST) or (request.GET.get("nuevo") == "1"),
    }
    return render(request, "rrhh/lista_empleados.html", context)


@rrhh_required
def detalle_empleado(request, pk):
    empleado = get_object_or_404(Empleado, pk=pk)
    documentos = empleado.documentos.all()
    ausencias = empleado.ausencias.all()
    form = SubirDocumentoForm()
    if request.method == "POST":
        form = SubirDocumentoForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.empleado = empleado
            doc.save()
            messages.success(request, "Documento subido correctamente.")
            return redirect("rrhh_detalle_empleado", pk=pk)
    return render(request, "rrhh/detalle_empleado.html", {
        "empleado": empleado, "documentos": documentos,
        "ausencias": ausencias, "form": form,
        "tipo_choices": Documento.TIPO_CHOICES,
    })


@rrhh_required
def editar_empleado(request, pk):
    empleado = get_object_or_404(Empleado, pk=pk)
    if request.method == "POST":
        form = EmpleadoForm(request.POST, request.FILES, instance=empleado)
        if form.is_valid():
            form.save()
            messages.success(request, f"Datos de {empleado.nombre} actualizados correctamente.")
            return redirect("rrhh_detalle_empleado", pk=pk)
    else:
        form = EmpleadoForm(instance=empleado)
    return render(request, "rrhh/editar_empleado.html", {"form": form, "empleado": empleado})


@rrhh_required
def editar_documento(request, pk):
    doc = get_object_or_404(Documento, pk=pk)
    if request.method == "POST":
        form = DocumentoEditForm(request.POST, instance=doc)
        if form.is_valid():
            form.save()
            messages.success(request, f"Documento '{doc.get_tipo_display()}' actualizado.")
        else:
            messages.error(request, "Error al actualizar el documento.")
    return redirect("rrhh_detalle_empleado", pk=doc.empleado.pk)


@rrhh_required
def eliminar_documento(request, pk):
    doc = get_object_or_404(Documento, pk=pk)
    empleado_pk = doc.empleado.pk
    if request.method == "POST":
        # Eliminar archivo físico
        if doc.archivo:
            if os.path.exists(doc.archivo.path):
                os.remove(doc.archivo.path)
        
        nombre = doc.get_tipo_display()
        doc.delete()
        messages.success(request, f"Documento '{nombre}' eliminado correctamente.")
    return redirect("rrhh_detalle_empleado", pk=empleado_pk)


@rrhh_required
def lista_ausencias(request):
    hoy = datetime.date.today()
    # Mostrar todas las ausencias activas (pendientes y en curso)
    ausencias = PeriodoAusencia.objects.select_related("empleado").filter(fecha_fin__gte=hoy).order_by("fecha_fin")
    # También mostrar ausencias históricas del mes actual
    ausencias_historial = PeriodoAusencia.objects.select_related("empleado").filter(
        fecha_fin__lt=hoy,
        fecha_inicio__year=hoy.year,
        fecha_inicio__month=hoy.month
    ).order_by("-fecha_inicio")[:20]

    from .forms import AusenciaForm
    ausencia_form = AusenciaForm()
    if request.method == "POST" and "registrar_ausencia" in request.POST:
        ausencia_form = AusenciaForm(request.POST)
        if ausencia_form.is_valid():
            ausencia_form.save()
            messages.success(request, "✅ Ausencia registrada correctamente.")
            return redirect("rrhh_ausencias")

    # Empleados en estado especial que quizás no tienen ausencia registrada
    emp_en_licencia = Empleado.objects.filter(estado="LICENCIA").count()
    emp_en_vacaciones = Empleado.objects.filter(estado="VACACIONES").count()

    return render(request, "rrhh/ausencias.html", {
        "ausencias": ausencias,
        "ausencias_historial": ausencias_historial,
        "hoy": hoy,
        "ausencia_form": ausencia_form,
        # Incluir TODOS los empleados (no solo ACTIVOS) para poder registrar ausencias también a quienes ya están en licencia/vacaciones
        "empleados": Empleado.objects.exclude(estado="FINIQUITADO").order_by("apellido", "nombre"),
        "show_modal": request.method == "POST" and "registrar_ausencia" in request.POST,
        "emp_en_licencia": emp_en_licencia,
        "emp_en_vacaciones": emp_en_vacaciones,
    })


@rrhh_required
def exportar_nomina_excel(request):
    """Exporta la nómina de empleados a Excel con formato profesional."""
    empleados = Empleado.objects.all().order_by("apellido", "nombre")
    
    hoy = datetime.date.today()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nómina Personal"

    # Estilos (consistentes con Inventario para ser profesional)
    color_oscuro = "0B1120"
    color_naranja = "F97316"
    color_gris = "64748B"
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    fill_titulo = PatternFill("solid", fgColor=color_oscuro)
    fill_header = PatternFill("solid", fgColor="1E293B")
    thin = Side(style="thin", color="374151")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # Encabezado Empresa
    ws.merge_cells("A1:G1")
    ws["A1"] = "CENTRO MÉDICO SAN LUCAS"
    ws["A1"].font = Font(name="Calibri", size=10, color=color_gris, italic=True)
    ws["A1"].fill = fill_titulo
    
    ws.merge_cells("A2:G2")
    ws["A2"] = "NÓMINA GENERAL DE PERSONAL"
    ws["A2"].font = Font(name="Calibri", size=16, bold=True, color=color_naranja)
    ws["A2"].fill = fill_titulo
    
    ws.merge_cells("A3:G3")
    ws["A3"] = f"Reporte generado el {hoy.strftime('%d/%m/%Y')}"
    ws["A3"].font = Font(name="Calibri", size=9, color="94A3B8")
    ws["A3"].fill = fill_titulo
    ws["A3"].alignment = left

    # Encabezados de tabla
    headers = [
        "APELLIDO", "NOMBRE", "RUT", "CARGO", "ÁREA", "ESTADO", 
        "F. NACIMIENTO", "NACIONALIDAD", "DIRECCIÓN",
        "F. INGRESO", "F. CONTRATO", "TURNO", "HORARIO",
        "EMERGENCIA (NOMBRE)", "EMERGENCIA (TEL)"
    ]
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = center
        cell.border = borde
        
    ws.row_dimensions[5].height = 20

    # Datos
    for i, emp in enumerate(empleados, 6):
        data = [
            emp.apellido, emp.nombre, emp.rut, emp.cargo,
            emp.get_area_display(), emp.get_estado_display(),
            emp.fecha_nacimiento.strftime("%d/%m/%Y") if emp.fecha_nacimiento else "—",
            emp.nacionalidad or "—",
            emp.direccion or "—",
            emp.fecha_ingreso.strftime("%d/%m/%Y") if emp.fecha_ingreso else "—",
            emp.fecha_contrato.strftime("%d/%m/%Y") if emp.fecha_contrato else "—",
            emp.turno or "—",
            emp.horario or "—",
            emp.emergencia_nombre or "—",
            emp.emergencia_telefono or "—"
        ]
        fill = PatternFill("solid", fgColor="111827" if i % 2 == 0 else "1F2937")
        
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(name="Calibri", size=9, color="F1F5F9")
            cell.fill = fill
            cell.border = borde
            cell.alignment = left if col <= 4 or col == 9 or col >= 14 else center
        ws.row_dimensions[i].height = 18

    # Ajustar columnas (A-O)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 35
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 15
    ws.column_dimensions["L"].width = 15
    ws.column_dimensions["M"].width = 20
    ws.column_dimensions["N"].width = 25
    ws.column_dimensions["O"].width = 15

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:O{len(empleados) + 5}"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Nomina_Personal_{hoy}.xlsx"'
    wb.save(response)
    return response

@rrhh_required
def lista_gastos_rrhh(request):
    """Muestra y registra gastos de RRHH (Pasajes, EPP, Caja Chica, etc.)."""
    hoy = datetime.date.today()
    gastos = GastoRRHH.objects.all().order_by("-fecha")
    
    # Filtros
    mes = request.GET.get("mes")
    anio = request.GET.get("anio")
    if not mes or not anio:
        mes = hoy.month
        anio = hoy.year
    
    gastos = gastos.filter(fecha__month=mes, fecha__year=anio)
    total_mes = gastos.aggregate(models.Sum('monto'))['monto__sum'] or 0
    
    form = GastoRRHHForm()
    if request.method == "POST":
        form = GastoRRHHForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Gasto registrado correctamente.")
            return redirect("rrhh_gastos")

    return render(request, "rrhh/gastos.html", {
        "gastos": gastos,
        "form": form,
        "total_mes": total_mes,
        "mes_actual": int(mes),
        "anio_actual": int(anio),
        "meses": range(1, 13),
        "anios": range(hoy.year - 2, hoy.year + 1),
    })

@rrhh_required
def exportar_gastos_excel(request):
    """Exporta el reporte mensual de gastos de RRHH a Excel."""
    mes = request.GET.get("mes")
    anio = request.GET.get("anio")
    
    if not mes or not anio:
        hoy = datetime.date.today()
        mes, anio = hoy.month, hoy.year
        
    gastos = GastoRRHH.objects.filter(fecha__month=mes, fecha__year=anio).order_by("fecha")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Gastos RRHH {mes}-{anio}"
    
    # Estilos Premium
    color_oscuro = "0B1120"
    color_verde = "10B981"
    fill_header = PatternFill("solid", fgColor="1E293B")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="374151")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Encabezados
    ws.merge_cells("A1:E1")
    ws["A1"] = f"REPORTE GASTOS RECURSOS HUMANOS - MES {mes}/{anio}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="F97316")
    ws["A1"].fill = PatternFill("solid", fgColor=color_oscuro)
    
    headers = ["FECHA", "CATEGORÍA", "DESCRIPCIÓN", "RESPONSABLE", "MONTO"]
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
        cell.border = borde
        
    # Datos
    total = 0
    for i, g in enumerate(gastos, 4):
        total += float(g.monto)
        data = [g.fecha.strftime("%d/%m/%Y"), g.get_categoria_display(), g.descripcion, g.responsable, float(g.monto)]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = borde
            if col == 5:
                cell.number_format = '"$"#,##0'
                cell.alignment = Alignment(horizontal="right")
                
    # Fila de Total
    row_total = len(gastos) + 4
    ws.merge_cells(f"A{row_total}:D{row_total}")
    ws[f"A{row_total}"] = "TOTAL GASTOS DEL MES"
    ws[f"A{row_total}"].font = Font(bold=True)
    ws[f"A{row_total}"].alignment = Alignment(horizontal="right")
    
    ws[f"E{row_total}"] = total
    ws[f"E{row_total}"].font = Font(bold=True, color=color_verde)
    ws[f"E{row_total}"].number_format = '"$"#,##0'
    ws[f"E{row_total}"].border = borde
    
    # Ajustar Columnas
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 18
    
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Gastos_RRHH_{mes}_{anio}.xlsx"'
    wb.save(response)
    return response

@rrhh_required
def inventario_epp(request):
    """Muestra el stock actual de artículos de RRHH."""
    productos = ProductoEPP.objects.all().order_by("nombre")
    # Últimas 10 entregas
    ultimas_entregas = EntregaEPP.objects.select_related("empleado", "producto").order_by("-fecha")[:15]
    
    entrega_form = EntregaEPPForm()
    
    return render(request, "rrhh/inventario_epp.html", {
        "productos": productos,
        "ultimas_entregas": ultimas_entregas,
        "entrega_form": entrega_form,
        "empleados": Empleado.objects.filter(estado="ACTIVO").order_by("apellido", "nombre"),
    })

@rrhh_required
def registrar_ingreso_epp(request):
    """Añadir stock a un producto de RRHH. Crea el producto si no existe."""
    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        talla = request.POST.get("talla", "").strip()
        cantidad_str = request.POST.get("cantidad", "0")
        precio_str = request.POST.get("precio_unitario", "")

        try:
            cantidad = float(cantidad_str)
        except (ValueError, TypeError):
            cantidad = 0

        if not nombre:
            messages.error(request, "Debes indicar el nombre del producto.")
            return redirect("rrhh_inventario_epp")

        if cantidad <= 0:
            messages.error(request, "La cantidad debe ser mayor a 0.")
            return redirect("rrhh_inventario_epp")

        # Buscar producto exacto por nombre+talla (case-insensitive) o crear nuevo
        prod = ProductoEPP.objects.filter(
            nombre__iexact=nombre, talla__iexact=talla
        ).first()

        if prod:
            prod.stock += cantidad
            if precio_str:
                try:
                    prod.precio_unitario = float(precio_str)
                except (ValueError, TypeError):
                    pass
            prod.save()
            messages.success(request, f"✅ {cantidad} unidades añadidas a {prod}.")
        else:
            # Crear nuevo producto EPP
            precio = 0
            if precio_str:
                try:
                    precio = float(precio_str)
                except (ValueError, TypeError):
                    precio = 0
            prod = ProductoEPP.objects.create(
                nombre=nombre,
                talla=talla,
                stock=cantidad,
                precio_unitario=precio,
            )
            messages.success(request, f"✅ Nuevo producto '{prod}' creado con {cantidad} unidades en stock.")

    return redirect("rrhh_inventario_epp")

@rrhh_required
def registrar_entrega_epp(request):
    """Registra la entrega de un implemento a un empleado y descuenta stock."""
    if request.method == "POST":
        form = EntregaEPPForm(request.POST)
        if form.is_valid():
            entrega = form.save(commit=False)
            producto = entrega.producto
            
            if producto.stock >= entrega.cantidad:
                producto.stock -= entrega.cantidad
                producto.save()
                entrega.save()
                
                # También registramos un GastoRRHH automático para que se vea en el Excel de gastos
                GastoRRHH.objects.create(
                    categoria="EPP",
                    fecha=datetime.date.today(),
                    monto=entrega.costo_total,
                    descripcion=f"ENTREGA: {entrega.cantidad} {producto.nombre}{' (T. '+producto.talla+')' if producto.talla else ''} a {entrega.empleado.get_full_name()}",
                    responsable=request.user.get_full_name()
                )
                
                messages.success(request, f"✅ Entrega de {producto.nombre} a {entrega.empleado} registrada correctamente.")
            else:
                messages.error(request, f"❌ Error: Stock insuficiente de {producto.nombre}. Disponible: {producto.stock}")
        else:
            messages.error(request, "❌ Error en el formulario de entrega.")
            
    return redirect("rrhh_inventario_epp")
