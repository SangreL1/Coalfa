import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\Coalfa\Desktop\Gestios_Citas_IS\Planilla de invntario 28-02-26.xlsx', data_only=True)
ws = wb.active

print("Filas donde CONCEPTO es None pero Articulo tiene valor (posibles subtotales):")
for i, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
    concepto, articulo, sku, um = row[0], row[1], row[2], row[3]
    if articulo and not concepto:
        print(f"  Fila {i}: concepto=None | articulo='{articulo}' | sku='{sku}' | um='{um}'")

print()
print("Filas donde SKU esta vacio y tiene articulo (20 muestras):")
count = 0
for i, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
    concepto, articulo, sku, um = row[0], row[1], row[2], row[3]
    if articulo and not sku and count < 20:
        print(f"  Fila {i}: concepto='{concepto}' | articulo='{articulo}' | sku='{sku}'")
        count += 1
